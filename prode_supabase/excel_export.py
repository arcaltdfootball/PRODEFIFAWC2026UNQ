import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ranking import obtener_ranking


def exportar_ranking():
    ranking = obtener_ranking()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranking FIFA 26"

    # ── Encabezados ──────────────────────────────────────────────────────────
    headers = ["Pos.", "Participante", "Puntos", "Aciertos", "Disputados", "Efectividad %"]
    header_fill = PatternFill("solid", fgColor="1a2a4a")
    header_font = Font(bold=True, color="FFD700", size=12)
    thin = Side(style="thin", color="2a3a5a")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[1].height = 22

    # ── Datos ─────────────────────────────────────────────────────────────────
    # obtener_ranking() devuelve: (nombre, puntos, prev_pos, aciertos, disputados)
    row_fill_odd  = PatternFill("solid", fgColor="0d1b2e")
    row_fill_even = PatternFill("solid", fgColor="112240")
    font_normal   = Font(color="FFFFFF", size=11)
    font_pos      = Font(color="FFD700", bold=True, size=13)

    for i, fila in enumerate(ranking, 1):
        nombre, puntos, prev_pos, aciertos, disputados = fila
        efectividad = round((aciertos / disputados) * 100, 1) if disputados > 0 else 0.0

        valores = [i, nombre, puntos, aciertos, disputados, efectividad]
        fill = row_fill_odd if i % 2 == 1 else row_fill_even

        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.fill = fill
            cell.font = font_pos if col == 1 else font_normal
            cell.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center"
            )
            cell.border = border

        ws.row_dimensions[i + 1].height = 20

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    # ── Guardar en memoria ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
